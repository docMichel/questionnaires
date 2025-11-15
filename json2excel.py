#!/usr/bin/env python3
"""
Convertit le JSON fusionné en tableau Excel
Usage: python json_to_excel.py fusion.json output.xlsx
"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def json_to_excel(json_file, excel_file):
    """Convertit JSON en Excel avec une ligne par page"""
    
    # Charger le JSON
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Créer workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Résultats"
    
    # === CONSTRUIRE L'EN-TÊTE ===
    headers = ['Page', 'Globale']
    
    # Récupérer toutes les questions possibles (de la première page complète)
    questions_template = {}
    for page in data['pages']:
        if page.get('questions'):
            for q_id, q_data in page['questions'].items():
                if q_id not in questions_template:
                    questions_template[q_id] = {
                        'titre': q_data.get('titre', q_id),
                        'reponses': [r['titre'] for r in q_data.get('reponses', [])]
                    }
    
    # Ajouter colonnes pour chaque question
    for q_id in sorted(questions_template.keys()):
        q_info = questions_template[q_id]
        # Ajouter une colonne par réponse possible
        for reponse_titre in q_info['reponses']:
            headers.append(f"{q_info['titre']} - {reponse_titre}")
    
    # Écrire l'en-tête
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # === REMPLIR LES DONNÉES ===
    for row_idx, page in enumerate(data['pages'], start=2):
        # Page et Globale
        ws.cell(row=row_idx, column=1, value=page['page'])
        ws.cell(row=row_idx, column=2, value=page.get('globale'))
        
        # Pour chaque question
        col_idx = 3
        for q_id in sorted(questions_template.keys()):
            q_info = questions_template[q_id]
            
            # Récupérer les données de cette page
            page_q_data = page.get('questions', {}).get(q_id, {})
            page_reponses = {r['titre']: r['cochee'] for r in page_q_data.get('reponses', [])}
            
            # Pour chaque réponse possible
            for reponse_titre in q_info['reponses']:
                cochee = page_reponses.get(reponse_titre, 0)
                cell = ws.cell(row=row_idx, column=col_idx, value=cochee)
                
                # Colorer si coché
                if cochee == 1:
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                
                col_idx += 1
    
    # Ajuster largeur des colonnes
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
    
    # Figer la première ligne
    ws.freeze_panes = "A2"
    
    # Sauver
    wb.save(excel_file)
    
    print(f"\n{'='*60}")
    print(f"CONVERSION JSON → EXCEL")
    print(f"{'='*60}\n")
    print(f"✓ Fichier source: {json_file}")
    print(f"✓ Fichier Excel: {excel_file}")
    print(f"  {len(data['pages'])} pages")
    print(f"  {len(questions_template)} questions")
    print(f"  {len(headers)} colonnes")
    print(f"\n{'='*60}\n")


def main():
    if len(sys.argv) < 3:
        print("\nUsage: python json_to_excel.py fusion.json output.xlsx\n")
        sys.exit(1)
    
    json_file = sys.argv[1]
    excel_file = sys.argv[2]
    
    json_to_excel(json_file, excel_file)


if __name__ == "__main__":
    main()